import hashlib
import datetime
import csv
import json
from typing import List, Dict, Optional
import openpyxl

##==== Foydali funksiyalar ====
def parol_xeshi(parol):
    return hashlib.sha256(parol.encode()).hexdigest()

def joriy_vaqtni_ol():
    return datetime.datetime.now().isoformat()

##==== Asosiy sinflar ====
class AbstraktRol:
    def __init__(self, _id, ism_familiya, email, parol):
        self._id = _id
        self._ism_familiya = ism_familiya
        self._email = email
        self._parol_xeshi = parol_xeshi(parol)
        self._yaratilgan = joriy_vaqtni_ol()

    def profilni_ol(self):
        raise NotImplementedError

    def profilni_yangila(self, ism_familiya=None, email=None):
        if ism_familiya:
            self._ism_familiya = ism_familiya
        if email:
            self._email = email

##==== Foydalanuvchi sinflari ====
class Foydalanuvchi(AbstraktRol):
    def __init__(self, _id, ism_familiya, email, parol, rol):
        super().__init__(_id, ism_familiya, email, parol)
        self.rol = rol
        self._xabarnomalar = []

    def xabar_qosh(self, matn):
        self._xabarnomalar.append({"id": len(self._xabarnomalar)+1, "matn": matn, "oqildi": False})

    def xabarlarni_korish(self):
        return self._xabarnomalar

    def xabarni_ochirish(self, xabar_id):
        self._xabarnomalar = [x for x in self._xabarnomalar if x["id"] != xabar_id]

class Oquvchi(Foydalanuvchi):
    def __init__(self, _id, ism_familiya, email, parol, sinf):
        super().__init__(_id, ism_familiya, email, parol, rol="Oquvchi")
        self.sinf = sinf
        self.fanlar = {}
        self.vazifalar = {}
        self.baholar = {}

    def vazifa_topshirish(self, vazifa_id, matn):
        self.vazifalar[vazifa_id] = {"matn": matn, "topshirilgan": joriy_vaqtni_ol()}

    def baholarni_korish(self, fan=None):
        return self.baholar.get(fan, []) if fan else self.baholar

    def ortacha_bahoni_hisobla(self):
        barcha_baholar = [b for baholar in self.baholar.values() for b in baholar]
        return sum(barcha_baholar) / len(barcha_baholar) if barcha_baholar else 0

class Oqituvchi(Foydalanuvchi):
    def __init__(self, _id, ism_familiya, email, parol):
        super().__init__(_id, ism_familiya, email, parol, rol="Oqituvchi")
        self.fanlar = []
        self.sinflar = []
        self.vazifalar = {}

    def vazifa_yarat(self, vid, sarlavha, tavsif, muddat, fan, sinf_id):
        vazifa = Vazifa(vid, sarlavha, tavsif, muddat, fan, self._id, sinf_id)
        self.vazifalar[vid] = vazifa
        return vazifa

    def vazifani_bahola(self, vazifa, oquvchi_id, baho):
        vazifa.baho_ber(oquvchi_id, baho)

    def oquvchi_yutuqlari(self, oquvchi):
        return oquvchi.baholarni_korish()

class Ota_ona(Foydalanuvchi):
    def __init__(self, _id, ism_familiya, email, parol):
        super().__init__(_id, ism_familiya, email, parol, rol="Otaxona")
        self.farzandlar = []

    def farzand_baholari(self, farzand):
        return farzand.baholarni_korish()

    def farzand_vazifalari(self, farzand):
        return farzand.vazifalar

class Admin(Foydalanuvchi):
    def __init__(self, _id, ism_familiya, email, parol):
        super().__init__(_id, ism_familiya, email, parol, rol="Admin")
        self.ruxsatlar = ["foydalanuvchi_qoshish", "foydalanuvchi_ochirish", "hisobot_yaratish"]

    def foydalanuvchi_qoshish(self, foydalanuvchi):
        return foydalanuvchi

    def foydalanuvchi_ochirish(self, royxat, user_id):
        return [f for f in royxat if f._id != user_id]

##==== Funksional sinflar ====
class Vazifa:
    def __init__(self, vid, sarlavha, tavsif, muddat, fan, oqituvchi_id, sinf_id):
        self.id = vid
        self.sarlavha = sarlavha
        self.tavsif = tavsif
        self.muddat = muddat
        self.fan = fan
        self.oqituvchi_id = oqituvchi_id
        self.sinf_id = sinf_id
        self.topshiriqlar = {}
        self.baholar = {}

    def topshiriq_qosh(self, oquvchi_id, matn):
        self.topshiriqlar[oquvchi_id] = matn

    def baho_ber(self, oquvchi_id, baho):
        self.baholar[oquvchi_id] = baho

    def holat(self):
        return {
            "topshirildi": len(self.topshiriqlar),
            "baholandi": len(self.baholar),
            "jami": len(set(self.topshiriqlar.keys()).union(set(self.baholar.keys())))
        }

class Baho:
    def __init__(self, bid, oquvchi_id, fan, qiymat, sana, oqituvchi_id):
        self.id = bid
        self.oquvchi_id = oquvchi_id
        self.fan = fan
        self.qiymat = qiymat
        self.sana = sana
        self.oqituvchi_id = oqituvchi_id

    def bahoni_yangila(self, qiymat):
        self.qiymat = qiymat

    def malumot(self):
        return self.__dict__

class Jadval:
    def __init__(self, jid, sinf_id, kun):
        self.id = jid
        self.sinf_id = sinf_id
        self.kun = kun
        self.darslar = {}

    def dars_qosh(self, vaqt, fan, oqituvchi_id):
        if vaqt in self.darslar:
            raise Exception("To'qnashuv aniqlangan")
        self.darslar[vaqt] = {"fan": fan, "oqituvchi_id": oqituvchi_id}

    def dars_ol(self, vaqt):
        self.darslar.pop(vaqt, None)

    def jadvalni_korish(self):
        return self.darslar

class Xabarnoma:
    def __init__(self, nid, matn, qabul_id):
        self.id = nid
        self.matn = matn
        self.qabul_id = qabul_id
        self.yaratilgan = joriy_vaqtni_ol()
        self.oqildi = False
        self.muhim = "oddiy"

    def yubor(self):
        return self.__dict__

    def belgilash_oqilgan(self):
        self.oqildi = True

##==== Eksport funksiyalari ====
def eksport_csvga(malumotlar: List[dict], fayl_nomi: str):
    if not malumotlar:
        return
    with open(fayl_nomi, mode='w', newline='', encoding='utf-8') as fayl:
        yozuvchi = csv.DictWriter(fayl, fieldnames=malumotlar[0].keys())
        yozuvchi.writeheader()
        yozuvchi.writerows(malumotlar)

def eksport_xlsxga(ma_dict: Dict[str, List[dict]], fayl_nomi: str):
    daftarcha = openpyxl.Workbook()
    for varaq_nomi, malumot in ma_dict.items():
        varaq = daftarcha.create_sheet(title=varaq_nomi)
        if not malumot:
            continue
        varaq.append(list(malumot[0].keys()))
        for qator in malumot:
            varaq.append(list(qator.values()))
    del daftarcha[daftarcha.sheetnames[0]]
    daftarcha.save(fayl_nomi)

def eksport_sqlga(malumotlar: List[dict], jadval_nomi: str):
    if not malumotlar:
        return ""
    ustunlar = malumotlar[0].keys()
    sql = f"CREATE TABLE {jadval_nomi} (" + ", ".join(f"{ustun} TEXT" for ustun in ustunlar) + ");\n"
    for qator in malumotlar:
        qiymatlar = [f"'{str(v)}'" for v in qator.values()]
        sql += f"INSERT INTO {jadval_nomi} (" + ", ".join(ustunlar) + ") VALUES (" + ", ".join(qiymatlar) + ");\n"
    return sql
